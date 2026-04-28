import os
import re
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.integrate import trapezoid


# ==================== File configuration ====================
# Indicate the number of sessions that occur every week (change as necessary).This is how the script bins and calculates variables calculated on a weekly basis.  
wk_session_num = 4

#Add participants and groups (these will be added to the summary file results). Make sure that the participant ID match the participant folder names exactly. 
GROUP_LOOKUP = {
    "MRS_01": "1",
    "MRS_02": "1",
    "MRS_03": "1",
    "MRS_04": "1",
    "MRS_05": "DO",
    "MRS_07": "2",
    "MRS_08": "2",
    "MRS_09": "2",
    "MRS_11": "1",
    "MRS_12": "WD",
    "MRS_13": "2",
    "MRS_14": "1",
    "MRS_15": "WD",
    "MRS_17": "2",
    "MRS_18": "1",
    "MRS_19": "1",
    "MRS_20": "DO",
    "MRS_22": "1",
    "MRS_24": "2",
    "MRS_25": "2",
    "MRS_26": "2",
    "MRS_27": "2",
    "MRS_28": "2",
    "MRS_29": "1",
    "MRS_30": "WD",
    "MRS_31": "1",
    "MRS_32": "2",
    "MRS_33": "2",
}

# ==================== Begin calculations ====================

# These constants set the airflow calculation values used throughout the script.
DIAMETER1 = 0.002  # 2 mm orifice for the smaller secondary hole
DIAMETER2 = 0      # Larger hole not used for inspiration
CD = 0.84          # Discharge coefficient for air at room temperature through a ~1.66 mm hole
RHO = 1.089416     # Air density at ambient conditions (kg/m^3)
AREA_ORIFICE = (np.pi * (DIAMETER1 / 2) ** 2) + (np.pi * (DIAMETER2 / 2) ** 2)

# Denote how you want the summary file saved
SUMMARY_PATH = Path("IMT_summary.xlsx")
DATA_DIR = Path("data")


def extract_trailing_number(filename):
    # Pull the session number off the end of a filename like MRS_07_12.
    match = re.search(r"_([0-9]+)$", Path(filename).stem)
    return int(match.group(1)) if match else None

def assign_bin(num):
    # Group every x sessions into one training week this is based on the  
    if num is None:
        return np.nan
    return ((num - 1) // wk_session_num) + 1

def lookup_group(study_id):
    # Match each participant ID to its study group.
    return GROUP_LOOKUP.get(study_id)


def bin_and_sum_by_number(df):
    # Add up weekly totals for work and SMIP.
    binned = df.groupby("Bin")[["Total_Work_J", "Total_SMIP"]].sum().reset_index()
    sigma_smip_per_bin = df.groupby("Bin")["Total_SMIP"].sum().tolist()
    return binned, sigma_smip_per_bin


def bin_and_max_by_number(df):
    # Find the highest total work and SMIP values reached within each week.
    binned = df.groupby("Bin")[["Total_Work_J", "Total_SMIP"]].max().reset_index()
    binned.columns = ["Bin", "Max_Total_Work_J", "Max_Total_SMIP"]
    return binned


def bin_and_max_breaths_by_number(df):
    # Find the strongest single-breath values within each week.
    breath_vars = ["SMIP_", "MIP_", "IT_", "Work_", "FIT_", "Max_inst_pwr_"]
    max_rows = []
    for _, group in df.groupby("Bin"):
        max_row = {}
        for var in breath_vars:
            breath_cols = [col for col in df.columns if col.startswith(var) and "Breath" in col]
            max_row[f"Max_{var}Breath"] = group[breath_cols].max().max() if breath_cols else np.nan
        max_rows.append(max_row)
    return pd.DataFrame(max_rows)


def load_or_create_workbook(summary_path):
    # Open the summary workbook if it exists, otherwise start a new one.
    if summary_path.exists():
        try:
            workbook = load_workbook(summary_path)
            if all(sheet.sheet_state != "visible" for sheet in workbook.worksheets):
                workbook.create_sheet("Sheet1")
            return workbook
        except Exception as exc:
            print(f"Error loading workbook: {exc}. Creating a new one.")

    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Sheet1"
    return workbook


def process_csv_file(filepath, bin_label):
    # Read one session file and calculate breath-by-breath metrics from the pressure trace.
    df = pd.read_csv(filepath, header=None, skiprows=6, usecols=[0, 1])
    df.columns = ["Time", "Pressure"]
    df["Time"] = pd.to_numeric(df["Time"], errors="coerce")
    df["Pressure"] = pd.to_numeric(df["Pressure"], errors="coerce")
    df["Pressure_Pa"] = df["Pressure"] * 98.0665

    nan_indices = list(df[df["Pressure"].isna()].index) + [len(df)]
    breath_data = {}
    total_work = 0
    total_smip = 0
    num_breaths = 0
    totals = {
        "time_above_30": 0,
        "time_above_50": 0,
        "time_above_80": 0,
        "intervention_time": 0,
    }
    week_totals = {"time_above_80": 0, "intervention_time": 0}
    first_breath_peak_pressure = None

    for i in range(len(nan_indices) - 1):
        # Split the file into separate breaths using blank rows in the export.
        start = nan_indices[i] + 1
        end = nan_indices[i + 1]
        segment = df.iloc[start:end].copy().dropna()
        if segment.empty:
            continue

        segment["Flow"] = CD * AREA_ORIFICE * np.sqrt(2 * np.abs(segment["Pressure_Pa"]) / RHO)
        segment["Power"] = segment["Flow"] * segment["Pressure_Pa"]

        breath_time = segment["Time"].max() - segment["Time"].min()
        totals["intervention_time"] += breath_time

        max_time = segment["Time"].max()
        max_pressure_pa = segment["Pressure_Pa"].max()
        max_pressure_cm_h2o = max_pressure_pa / 98.0665
        if first_breath_peak_pressure is None:
            first_breath_peak_pressure = max_pressure_cm_h2o
        max_power = segment["Power"].max()
        work = trapezoid(segment["Power"].values, segment["Time"].values)
        smip = trapezoid(segment["Pressure"].values, segment["Time"].values)

        # Track how long the breath stays above thresholds based on the first breath's peak pressure in the file.
        for pct, key in [(0.3, "time_above_30"), (0.5, "time_above_50"), (0.8, "time_above_80")]:
            threshold = pct * first_breath_peak_pressure
            segment["Above"] = segment["Pressure"] >= threshold
            time_above = segment.loc[segment["Above"], "Time"].diff().fillna(0).sum()
            totals[key] += time_above

            if pct == 0.8:
                week_totals["time_above_80"] += time_above
                week_totals["intervention_time"] += breath_time

        # Estimate FIT by looking at the work done up to the first 500 mL of inspired volume.
        flow_vals = segment["Flow"].values * 1000  # L/s
        time_vals = segment["Time"].values
        delta_t = np.diff(time_vals, prepend=time_vals[0])
        cumulative_volume = np.cumsum(delta_t * flow_vals)
        cutoff_index = np.argmax(cumulative_volume >= 0.5)

        if cumulative_volume[cutoff_index] < 0.5:
            sip_500ml = np.nan
            fit = np.nan
        else:
            sip_segment = segment.iloc[: cutoff_index + 1]
            sip_500ml = trapezoid(sip_segment["Pressure"].values, sip_segment["Time"].values)
            if sip_500ml == 0:
                fit = np.nan
            else:
                time_500ml = segment.iloc[cutoff_index]["Time"]
                fit = (smip * max_time) / (sip_500ml * time_500ml) if time_500ml else np.nan

        total_work += work
        total_smip += smip
        num_breaths += 1

        breath_label = f"Breath_{i + 1}"
        breath_data[f"IT_{breath_label}_s"] = max_time
        breath_data[f"MIP_{breath_label}_cm_H2O"] = max_pressure_cm_h2o
        breath_data[f"Work_{breath_label}_J"] = work
        breath_data[f"SMIP_{breath_label}_cmH20_s"] = smip
        breath_data[f"Max_inst_pwr_{breath_label}_W"] = max_power
        breath_data[f"FIT_{breath_label}"] = fit

    avg_work = total_work / num_breaths if num_breaths > 0 else 0
    row_data = {
        "File": filepath.stem,
        "File_Number": extract_trailing_number(filepath.name),
        "Bin": bin_label,
        "Total_Work_J": total_work,
        "Total_SMIP": total_smip,
        "Average_Work_per_breath_J": avg_work,
    }
    row_data.update(breath_data)

    return row_data, totals, week_totals


def process_participant_folder(folder_path):
    # Process one participant folder so multiple participants can be handled in parallel.
    folder_path = Path(folder_path)
    folder = folder_path.name
    summary_rows = []
    total_subfolder_work = 0
    total_subfolder_smip = 0
    total_time_above_30 = 0
    total_time_above_50 = 0
    total_time_above_80 = 0
    total_intervention_time = 0
    time_above_80_per_week = {}
    intervention_time_per_week = {}

    for filename in sorted(os.listdir(folder_path)):
        # Process each session file and add it to the participant summary.
        if not filename.endswith(".csv"):
            continue

        trailing_number = extract_trailing_number(filename)
        bin_label = assign_bin(trailing_number)
        filepath = folder_path / filename

        try:
            row_data, totals, week_totals = process_csv_file(filepath, bin_label)
            summary_rows.append(row_data)

            total_subfolder_work += row_data["Total_Work_J"]
            total_subfolder_smip += row_data["Total_SMIP"]
            total_time_above_30 += totals["time_above_30"]
            total_time_above_50 += totals["time_above_50"]
            total_time_above_80 += totals["time_above_80"]
            total_intervention_time += totals["intervention_time"]

            time_above_80_per_week[bin_label] = (
                time_above_80_per_week.get(bin_label, 0) + week_totals["time_above_80"]
            )
            intervention_time_per_week[bin_label] = (
                intervention_time_per_week.get(bin_label, 0) + week_totals["intervention_time"]
            )
        except Exception as exc:
            print(f"Error processing file {filename}: {exc}")

    if not summary_rows:
        return None

    summary_df = pd.DataFrame(summary_rows)
    binned_averages, sigma_smip_per_bin = bin_and_sum_by_number(summary_df)
    breath_max_per_bin = bin_and_max_breaths_by_number(summary_df)
    weekly_maxes = bin_and_max_by_number(summary_df)

    # Combine weekly totals, weekly peaks, and peak single-breath values into one table.
    binned_averages.rename(columns={"Bin": "Week"}, inplace=True)
    weekly_maxes.rename(columns={"Bin": "Week"}, inplace=True)

    binned_averages.insert(0, "Subject_Folder", [folder] * len(binned_averages))
    binned_averages.insert(1, "Group", [lookup_group(folder)] * len(binned_averages))
    binned_averages.insert(
        4,
        "Total_intervention_IM_Work_J",
        [total_subfolder_work] + [None] * (len(binned_averages) - 1),
    )
    binned_averages.insert(
        5,
        "Total_intervention_SMIP",
        [total_subfolder_smip] + [None] * (len(binned_averages) - 1),
    )
    binned_averages.insert(
        6,
        "sigma_SMIP_wk",
        sigma_smip_per_bin + [None] * (len(binned_averages) - len(sigma_smip_per_bin)),
    )
    binned_averages.insert(
        7,
        "Time_above_30_pct",
        [total_time_above_30] + [None] * (len(binned_averages) - 1),
    )
    binned_averages.insert(
        8,
        "Time_above_50_pct",
        [total_time_above_50] + [None] * (len(binned_averages) - 1),
    )
    binned_averages.insert(
        9,
        "Time_above_80_pct",
        [total_time_above_80] + [None] * (len(binned_averages) - 1),
    )
    binned_averages.insert(
        10,
        "Total_intervention_time",
        [total_intervention_time] + [None] * (len(binned_averages) - 1),
    )
    binned_averages.insert(
        11,
        "Pct_time_above_30_pct",
        [((total_time_above_30 / total_intervention_time) * 100) if total_intervention_time > 0 else None]
        + [None] * (len(binned_averages) - 1),
    )
    binned_averages.insert(
        12,
        "Pct_time_above_50_pct",
        [((total_time_above_50 / total_intervention_time) * 100) if total_intervention_time > 0 else None]
        + [None] * (len(binned_averages) - 1),
    )
    binned_averages.insert(
        13,
        "Pct_time_above_80_pct",
        [((total_time_above_80 / total_intervention_time) * 100) if total_intervention_time > 0 else None]
        + [None] * (len(binned_averages) - 1),
    )
    binned_averages["Pct_time_above_80_per_week"] = binned_averages["Week"].apply(
        lambda wk: (time_above_80_per_week.get(wk, 0) / intervention_time_per_week.get(wk, 1) * 100)
    )

    combined = pd.concat(
        [
            binned_averages.reset_index(drop=True),
            weekly_maxes.reset_index(drop=True).drop(columns="Week"),
            breath_max_per_bin.reset_index(drop=True),
        ],
        axis=1,
    )
    combined = combined.loc[:, ~combined.columns.duplicated()]

    return {
        "folder": folder,
        "summary_df": summary_df,
        "combined": combined,
    }


def build_weekly_summary(summary_path=SUMMARY_PATH, data_dir=DATA_DIR):
    # Walk through every participant folder, process them in parallel, and build the workbook summary tabs.
    workbook = load_or_create_workbook(summary_path)
    imt_summary_rows = []
    folder_paths = sorted([folder_path for folder_path in data_dir.iterdir() if folder_path.is_dir()])
    max_workers = os.cpu_count() or 1
    participant_results = []
    print(f"Processing participant folders using {max_workers} CPU worker(s).")

    # Use all available CPU cores to process participant folders at the same time.
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_participant_folder, folder_path): folder_path.name for folder_path in folder_paths}
        for future in as_completed(futures):
            folder_name = futures[future]
            try:
                result = future.result()
                if result is not None:
                    participant_results.append(result)
            except Exception as exc:
                print(f"Error processing participant folder {folder_name}: {exc}")

    for result in sorted(participant_results, key=lambda item: item["folder"]):
        imt_summary_rows.append(result["combined"])

        if result["folder"] in workbook.sheetnames:
            workbook.remove(workbook[result["folder"]])
        sheet = workbook.create_sheet(result["folder"])

        # Save the raw per-session results for this participant to its own tab.
        for row in dataframe_to_rows(result["summary_df"], index=False, header=True):
            sheet.append(row)

    if not imt_summary_rows:
        raise ValueError(f"No participant CSV files were processed from '{data_dir}'.")

    imt_summary_df = pd.concat(imt_summary_rows, ignore_index=True)

    if "Weekly Summaries" in workbook.sheetnames:
        workbook.remove(workbook["Weekly Summaries"])
    weekly_sheet = workbook.create_sheet("Weekly Summaries")

    # Save the combined weekly summary across all participants.
    for row in dataframe_to_rows(imt_summary_df, index=False, header=True):
        weekly_sheet.append(row)

    workbook._sheets.insert(0, workbook._sheets.pop(workbook.sheetnames.index("Weekly Summaries")))

    if "Sheet" in workbook.sheetnames and workbook["Sheet"].max_row == 1 and workbook["Sheet"].max_column == 1:
        workbook.remove(workbook["Sheet"])

    workbook.save(summary_path)
    print("Weekly summaries created and saved to IMT_summary.xlsx.")


def process_and_add_to_spss_summary(summary_path=SUMMARY_PATH):
    # Convert the weekly summary into a wide participant-level layout for SPSS.
    if not summary_path.exists():
        raise FileNotFoundError(f"The file '{summary_path}' does not exist in the current working directory.")

    df = pd.read_excel(summary_path, sheet_name="Weekly Summaries")
    per_week_cols = [
        "Total_Work_J",
        "sigma_SMIP_wk",
        "Max_Total_Work_J",
        "Max_Total_SMIP",
        "Pct_time_above_80_per_week",
    ]
    breath_cols = [col for col in df.columns if "Max_" in col and "Breath" in col and col not in per_week_cols]
    per_week_cols += breath_cols

    # Build one row per participant with separate columns for each week.
    all_weeks = sorted(df["Week"].dropna().unique())
    header = [
        "Participant",
        "Group",
        "Total_intervention_IM_Work_J",
        "Total_intervention_SMIP",
        "Total_intervention_time",
        "Time_above_30_pct",
        "Pct_time_above_30",
        "Time_above_50_pct",
        "Pct_time_above_50",
        "Time_above_80_pct",
        "Pct_time_above_80",
    ]
    for week_num in all_weeks:
        for col in per_week_cols:
            header.append(f"{col}_wk_{int(week_num)}")

    spss_rows = []
    for participant in df["Subject_Folder"].unique():
        part_df = df[df["Subject_Folder"] == participant]
        group = part_df["Group"].iloc[0] if "Group" in part_df.columns else lookup_group(participant)
        total_work = part_df["Total_intervention_IM_Work_J"].iloc[0]
        total_smip = part_df["Total_intervention_SMIP"].iloc[0]
        total_time = part_df["Total_intervention_time"].iloc[0]
        time30 = part_df["Time_above_30_pct"].iloc[0]
        time50 = part_df["Time_above_50_pct"].iloc[0]
        time80 = part_df["Time_above_80_pct"].iloc[0]
        pct30 = (time30 / total_time * 100) if total_time else None
        pct50 = (time50 / total_time * 100) if total_time else None
        pct80 = (time80 / total_time * 100) if total_time else None

        row = [participant, group, total_work, total_smip, total_time, time30, pct30, time50, pct50, time80, pct80]

        for week_num in all_weeks:
            week_row = part_df[part_df["Week"] == week_num]
            if not week_row.empty:
                for col in per_week_cols:
                    row.append(week_row.iloc[0][col])
            else:
                row.extend([None] * len(per_week_cols))

        spss_rows.append(row)

    workbook = load_workbook(summary_path)
    if "IMT Summary SPSS" in workbook.sheetnames:
        workbook.remove(workbook["IMT Summary SPSS"])
    spss_sheet = workbook.create_sheet("IMT Summary SPSS")

    # Write the wide-format SPSS table back into the workbook.
    spss_sheet.append(header)
    for row in spss_rows:
        spss_sheet.append(row)

    sheet_order = workbook.sheetnames
    sheet_order.remove("IMT Summary SPSS")
    sheet_order.insert(1, "IMT Summary SPSS")
    workbook._sheets = [workbook[sheet] for sheet in sheet_order]
    workbook.save(summary_path)
    print("SPSS summary sheet updated in IMT_summary.xlsx.")


def create_long_format_summary(summary_path=SUMMARY_PATH):
    # Turn the wide SPSS sheet back into a long table with one row per participant-week.
    wide_sheet_name = "IMT Summary SPSS"
    long_sheet_name = "IMST summary long format"
    df = pd.read_excel(summary_path, sheet_name=wide_sheet_name)

    static_cols = [
        "Participant",
        "Group",
        "Total_intervention_IM_Work_J",
        "Total_intervention_SMIP",
        "Total_intervention_time",
        "Time_above_30_pct",
        "Pct_time_above_30",
        "Time_above_50_pct",
        "Pct_time_above_50",
        "Time_above_80_pct",
        "Pct_time_above_80",
    ]

    week_column_map = {}
    for col in df.columns:
        # Match columns like Total_Work_J_wk_3 so they can be unpacked by week.
        match = re.match(r"(.+)_wk_(\d+)$", str(col))
        if match:
            base_name, week_num = match.groups()
            week_column_map.setdefault(int(week_num), {})[base_name] = col

    long_rows = []
    for _, participant_row in df.iterrows():
        static_values = {col: participant_row.get(col) for col in static_cols}

        for week_num in sorted(week_column_map):
            # Build one long-format row using the participant info plus that week's values.
            row = {"Week": week_num}
            row.update(static_values)

            has_week_data = False
            for base_name, source_col in week_column_map[week_num].items():
                value = participant_row.get(source_col)
                row[base_name] = value
                if pd.notna(value):
                    has_week_data = True

            if has_week_data:
                long_rows.append(row)

    long_df = pd.DataFrame(long_rows)
    if not long_df.empty:
        ordered_week_cols = sorted({base_name for week_cols in week_column_map.values() for base_name in week_cols})
        ordered_cols = ["Participant", "Week"] + [col for col in static_cols if col != "Participant"] + ordered_week_cols
        long_df = long_df.reindex(columns=ordered_cols)

    workbook = load_workbook(summary_path)
    if long_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[long_sheet_name])
    long_sheet = workbook.create_sheet(long_sheet_name)

    # Save the long-format table as its own workbook tab.
    for row in dataframe_to_rows(long_df, index=False, header=True):
        long_sheet.append(row)

    sheet_order = workbook.sheetnames
    sheet_order.remove(long_sheet_name)
    insert_at = 2 if len(sheet_order) >= 2 else len(sheet_order)
    sheet_order.insert(insert_at, long_sheet_name)
    workbook._sheets = [workbook[sheet] for sheet in sheet_order]
    workbook.save(summary_path)
    print("Long-format summary sheet updated in IMT_summary.xlsx.")


def main():
    # Run the full pipeline: session processing, wide summary, then long summary.
    build_weekly_summary()
    process_and_add_to_spss_summary()
    create_long_format_summary()
    print("IMST processing complete.")


if __name__ == "__main__":
    main()
